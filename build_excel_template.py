"""
build_excel_template.py
--------------------------
Generates a professional, formula-driven Excel DCF Model
(data/DCF_Model_Template.xlsx).

This is the "Excel DCF basics" version of the model — useful for
sharing with people who don't use Python, or as the first thing you
build before moving to the Python/Streamlit version.

Run with:
    python build_excel_template.py

Sheets created:
    1. Assumptions   - all editable inputs (blue text, yellow fill)
    2. WACC           - CAPM + WACC calculation
    3. Forecast       - 5-year Revenue -> EBIT -> FCFF build-up
    4. DCF Valuation  - Enterprise Value -> Equity Value -> Value/Share
    5. Sensitivity    - WACC vs Terminal Growth grid
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# STYLE CONSTANTS  (per financial-modeling convention)
# ----------------------------------------------------------------------
FONT_NAME = "Arial"

INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=11)              # blue = hardcoded input
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=11)            # black = formula
LINK_FONT = Font(name=FONT_NAME, color="008000", size=11)               # green = link to another sheet
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", size=12, bold=True)
TITLE_FONT = Font(name=FONT_NAME, color="FFFFFF", size=16, bold=True)
LABEL_FONT = Font(name=FONT_NAME, color="000000", size=11, bold=False)
BOLD_LABEL_FONT = Font(name=FONT_NAME, color="000000", size=11, bold=True)
RESULT_FONT = Font(name=FONT_NAME, color="000000", size=13, bold=True)

HEADER_FILL = PatternFill("solid", fgColor="1F2C3A")
KEY_ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")
RESULT_FILL = PatternFill("solid", fgColor="D9F2E6")
TITLE_FILL = PatternFill("solid", fgColor="0F2027")

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = "#,##0;(#,##0)"
PCT_FMT = "0.00%"
NUM_FMT_2DP = "#,##0.00;(#,##0.00)"


def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    top_left = cell_range.split(":")[0]
    ws[top_left] = text
    ws[top_left].font = TITLE_FONT
    ws[top_left].fill = TITLE_FILL
    ws[top_left].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[int(top_left[1:])].height = 30


def style_section_header(ws, cell, text):
    ws[cell] = text
    ws[cell].font = HEADER_FONT
    ws[cell].fill = HEADER_FILL
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)


def label(ws, cell, text, bold=False):
    ws[cell] = text
    ws[cell].font = BOLD_LABEL_FONT if bold else LABEL_FONT


def input_cell(ws, cell, value, number_format=None, key_assumption=False):
    ws[cell] = value
    ws[cell].font = INPUT_FONT
    ws[cell].border = BORDER
    if number_format:
        ws[cell].number_format = number_format
    if key_assumption:
        ws[cell].fill = KEY_ASSUMPTION_FILL


def formula_cell(ws, cell, formula, number_format=None, bold=False, result=False):
    ws[cell] = formula
    ws[cell].font = RESULT_FONT if result else FORMULA_FONT
    ws[cell].border = BORDER
    if number_format:
        ws[cell].number_format = number_format
    if result:
        ws[cell].fill = RESULT_FILL


def build_workbook(output_path="data/DCF_Model_Template.xlsx"):
    wb = openpyxl.Workbook()

    # ================================================================
    # SHEET 1: ASSUMPTIONS
    # ================================================================
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45

    style_title(ws, "A1:C1", "DCF Valuation Model - Assumptions")

    style_section_header(ws, "A3", "Company Info")
    label(ws, "A4", "Company Name")
    input_cell(ws, "B4", "TCS")
    label(ws, "A5", "Current Market Price (Rs)")
    input_cell(ws, "B5", 3200, CURRENCY_FMT)

    style_section_header(ws, "A7", "Base Financials (Latest Actual Year)")
    label(ws, "A8", "Base Revenue (Rs Cr)")
    input_cell(ws, "B8", 240893, CURRENCY_FMT, key_assumption=True)
    label(ws, "A9", "Net Debt (Rs Cr)  [negative = net cash]")
    input_cell(ws, "B9", -22400, CURRENCY_FMT, key_assumption=True)
    label(ws, "A10", "Shares Outstanding (Cr)")
    input_cell(ws, "B10", 365, NUM_FMT_2DP, key_assumption=True)

    style_section_header(ws, "A12", "Forecast Assumptions")
    label(ws, "A13", "Revenue Growth Rate (%)")
    input_cell(ws, "B13", 0.08, PCT_FMT, key_assumption=True)
    label(ws, "A14", "EBIT Margin (%)")
    input_cell(ws, "B14", 0.26, PCT_FMT, key_assumption=True)
    label(ws, "A15", "Tax Rate (%)")
    input_cell(ws, "B15", 0.25, PCT_FMT, key_assumption=True)
    label(ws, "A16", "D&A (% of Revenue)")
    input_cell(ws, "B16", 0.03, PCT_FMT)
    label(ws, "A17", "Capex (% of Revenue)")
    input_cell(ws, "B17", 0.05, PCT_FMT)
    label(ws, "A18", "Change in NWC (% of Revenue)")
    input_cell(ws, "B18", 0.01, PCT_FMT)
    label(ws, "A19", "Forecast Years")
    input_cell(ws, "B19", 5, "0")

    style_section_header(ws, "A21", "WACC Inputs")
    label(ws, "A22", "Risk Free Rate (%)")
    input_cell(ws, "B22", 0.068, PCT_FMT, key_assumption=True)
    label(ws, "A23", "Beta")
    input_cell(ws, "B23", 0.95, NUM_FMT_2DP, key_assumption=True)
    label(ws, "A24", "Expected Market Return (%)")
    input_cell(ws, "B24", 0.115, PCT_FMT, key_assumption=True)
    label(ws, "A25", "Pre-tax Cost of Debt (%)")
    input_cell(ws, "B25", 0.075, PCT_FMT)
    label(ws, "A26", "Market Capitalization (Rs Cr)")
    input_cell(ws, "B26", 1160000, CURRENCY_FMT)
    label(ws, "A27", "Total Debt (Rs Cr)")
    input_cell(ws, "B27", 8000, CURRENCY_FMT)

    style_section_header(ws, "A29", "Terminal Value")
    label(ws, "A30", "Terminal Growth Rate (%)")
    input_cell(ws, "B30", 0.04, PCT_FMT, key_assumption=True)

    ws["A32"] = "Yellow cells = key assumptions you should adjust. Blue text = editable input."
    ws["A32"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")
    ws.merge_cells("A32:C32")

    # ================================================================
    # SHEET 2: WACC
    # ================================================================
    ws2 = wb.create_sheet("WACC")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 45

    style_title(ws2, "A1:C1", "WACC - Discount Rate Calculation")

    style_section_header(ws2, "A3", "Cost of Equity (CAPM)")
    label(ws2, "A4", "Risk Free Rate")
    formula_cell(ws2, "B4", "=Assumptions!B22", PCT_FMT)
    ws2["B4"].font = LINK_FONT
    label(ws2, "A5", "Beta")
    formula_cell(ws2, "B5", "=Assumptions!B23", NUM_FMT_2DP)
    ws2["B5"].font = LINK_FONT
    label(ws2, "A6", "Expected Market Return")
    formula_cell(ws2, "B6", "=Assumptions!B24", PCT_FMT)
    ws2["B6"].font = LINK_FONT
    label(ws2, "A7", "Cost of Equity (Ke)", bold=True)
    formula_cell(ws2, "B7", "=B4+B5*(B6-B4)", PCT_FMT, result=True)

    style_section_header(ws2, "A9", "Cost of Debt")
    label(ws2, "A10", "Pre-tax Cost of Debt")
    formula_cell(ws2, "B10", "=Assumptions!B25", PCT_FMT)
    ws2["B10"].font = LINK_FONT
    label(ws2, "A11", "Tax Rate")
    formula_cell(ws2, "B11", "=Assumptions!B15", PCT_FMT)
    ws2["B11"].font = LINK_FONT
    label(ws2, "A12", "After-tax Cost of Debt", bold=True)
    formula_cell(ws2, "B12", "=B10*(1-B11)", PCT_FMT, result=True)

    style_section_header(ws2, "A14", "Capital Structure Weights")
    label(ws2, "A15", "Market Cap (Equity, E)")
    formula_cell(ws2, "B15", "=Assumptions!B26", CURRENCY_FMT)
    ws2["B15"].font = LINK_FONT
    label(ws2, "A16", "Total Debt (D)")
    formula_cell(ws2, "B16", "=Assumptions!B27", CURRENCY_FMT)
    ws2["B16"].font = LINK_FONT
    label(ws2, "A17", "Equity Weight  E/(D+E)")
    formula_cell(ws2, "B17", "=B15/(B15+B16)", PCT_FMT)
    label(ws2, "A18", "Debt Weight  D/(D+E)")
    formula_cell(ws2, "B18", "=B16/(B15+B16)", PCT_FMT)

    style_section_header(ws2, "A20", "Final WACC")
    label(ws2, "A21", "WACC = E/(D+E) x Ke + D/(D+E) x Kd x (1-T)", bold=True)
    formula_cell(ws2, "B21", "=B17*B7+B18*B12", PCT_FMT, result=True)

    # ================================================================
    # SHEET 3: FORECAST
    # ================================================================
    ws3 = wb.create_sheet("Forecast")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 28
    for col in "BCDEFG":
        ws3.column_dimensions[col].width = 15

    style_title(ws3, "A1:G1", "5-Year Forecast - Revenue to FCFF")

    style_section_header(ws3, "A3", "Period")
    headers = ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for i, h in enumerate(headers):
        cell = f"{get_column_letter(i + 1)}4"
        ws3[cell] = h
        ws3[cell].font = BOLD_LABEL_FONT
        ws3[cell].fill = PatternFill("solid", fgColor="E8EEF4")
        ws3[cell].border = BORDER

    row_labels = [
        ("Revenue", 5), ("EBIT", 6), ("NOPAT", 7),
        ("(+) D&A", 8), ("(-) Capex", 9), ("(-) Change in NWC", 10),
        ("FCFF", 11), ("Discount Factor", 12), ("PV of FCFF", 13),
    ]
    for text, r in row_labels:
        bold = text in ("FCFF", "PV of FCFF")
        label(ws3, f"A{r}", text, bold=bold)

    cols = ["B", "C", "D", "E", "F"]
    for i, c in enumerate(cols):
        if i == 0:
            formula_cell(ws3, f"{c}5", "=Assumptions!$B$8*(1+Assumptions!$B$13)", CURRENCY_FMT)
        else:
            prev = cols[i - 1]
            formula_cell(ws3, f"{c}5", f"={prev}5*(1+Assumptions!$B$13)", CURRENCY_FMT)

        formula_cell(ws3, f"{c}6", f"={c}5*Assumptions!$B$14", CURRENCY_FMT)
        formula_cell(ws3, f"{c}7", f"={c}6*(1-Assumptions!$B$15)", CURRENCY_FMT)
        formula_cell(ws3, f"{c}8", f"={c}5*Assumptions!$B$16", CURRENCY_FMT)
        formula_cell(ws3, f"{c}9", f"={c}5*Assumptions!$B$17", CURRENCY_FMT)
        formula_cell(ws3, f"{c}10", f"={c}5*Assumptions!$B$18", CURRENCY_FMT)
        formula_cell(ws3, f"{c}11", f"={c}7+{c}8-{c}9-{c}10", CURRENCY_FMT, result=True)
        formula_cell(ws3, f"{c}12", f"=1/(1+WACC!$B$21)^{i + 1}", "0.0000")
        formula_cell(ws3, f"{c}13", f"={c}11*{c}12", CURRENCY_FMT, result=True)

    # ================================================================
    # SHEET 4: DCF VALUATION
    # ================================================================
    ws4 = wb.create_sheet("DCF Valuation")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 40

    style_title(ws4, "A1:C1", "DCF Valuation Summary")

    style_section_header(ws4, "A3", "Present Value of Cash Flows")
    label(ws4, "A4", "Sum of PV of FCFF (Years 1-5)", bold=True)
    formula_cell(ws4, "B4", "=SUM(Forecast!B13:F13)", CURRENCY_FMT, result=True)

    style_section_header(ws4, "A6", "Terminal Value")
    label(ws4, "A7", "Final Year FCFF")
    formula_cell(ws4, "B7", "=Forecast!F11", CURRENCY_FMT)
    ws4["B7"].font = LINK_FONT
    label(ws4, "A8", "Terminal Growth Rate")
    formula_cell(ws4, "B8", "=Assumptions!B30", PCT_FMT)
    ws4["B8"].font = LINK_FONT
    label(ws4, "A9", "WACC")
    formula_cell(ws4, "B9", "=WACC!B21", PCT_FMT)
    ws4["B9"].font = LINK_FONT
    label(ws4, "A10", "Terminal Value = FCFF5 x (1+g) / (WACC-g)", bold=True)
    formula_cell(ws4, "B10", "=B7*(1+B8)/(B9-B8)", CURRENCY_FMT, result=True)
    label(ws4, "A11", "PV of Terminal Value", bold=True)
    formula_cell(ws4, "B11", "=B10/(1+B9)^Assumptions!B19", CURRENCY_FMT, result=True)

    style_section_header(ws4, "A13", "Enterprise Value -> Equity Value")
    label(ws4, "A14", "Enterprise Value", bold=True)
    formula_cell(ws4, "B14", "=B4+B11", CURRENCY_FMT, result=True)
    label(ws4, "A15", "(-) Net Debt")
    formula_cell(ws4, "B15", "=Assumptions!B9", CURRENCY_FMT)
    ws4["B15"].font = LINK_FONT
    label(ws4, "A16", "Equity Value", bold=True)
    formula_cell(ws4, "B16", "=B14-B15", CURRENCY_FMT, result=True)
    label(ws4, "A17", "/ Shares Outstanding")
    formula_cell(ws4, "B17", "=Assumptions!B10", NUM_FMT_2DP)
    ws4["B17"].font = LINK_FONT

    style_section_header(ws4, "A19", "Intrinsic Value")
    label(ws4, "A20", "Intrinsic Value per Share (Rs)", bold=True)
    formula_cell(ws4, "B20", "=B16/B17", CURRENCY_FMT, result=True)
    ws4["B20"].font = Font(name=FONT_NAME, size=15, bold=True, color="0F5132")

    label(ws4, "A21", "Current Market Price (Rs)")
    formula_cell(ws4, "B21", "=Assumptions!B5", CURRENCY_FMT)
    ws4["B21"].font = LINK_FONT
    label(ws4, "A22", "Upside / (Downside) %", bold=True)
    formula_cell(ws4, "B22", "=(B20-B21)/B21", PCT_FMT, result=True)

    label(ws4, "A24", "Verdict", bold=True)
    formula_cell(ws4, "B24", '=IF(B22>0,"Potentially Undervalued","Potentially Overvalued")', bold=True, result=True)

    ws4["A26"] = "Not a guaranteed Buy/Sell signal - DCF output is highly sensitive to assumptions. See the Sensitivity sheet."
    ws4["A26"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")
    ws4.merge_cells("A26:C26")

    # ================================================================
    # SHEET 5: SENSITIVITY (WACC vs Terminal Growth)
    # ================================================================
    ws5 = wb.create_sheet("Sensitivity")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 20
    for col in "BCDEF":
        ws5.column_dimensions[col].width = 14

    style_title(ws5, "A1:F1", "Sensitivity - Intrinsic Value per Share")
    ws5["A3"] = "Rows = WACC, Columns = Terminal Growth Rate"
    ws5["A3"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")

    ws5["A4"] = "WACC \\ Growth"
    ws5["A4"].font = BOLD_LABEL_FONT
    ws5["A4"].fill = PatternFill("solid", fgColor="E8EEF4")
    ws5["A4"].border = BORDER

    growth_offsets = [-0.01, -0.005, 0, 0.005, 0.01]
    wacc_offsets = [-0.01, -0.005, 0, 0.005, 0.01]
    growth_cols = ["B", "C", "D", "E", "F"]
    wacc_rows = [5, 6, 7, 8, 9]

    for i, c in enumerate(growth_cols):
        formula_cell(ws5, f"{c}4", f"=Assumptions!$B$30+({growth_offsets[i]})", PCT_FMT)
        ws5[f"{c}4"].font = BOLD_LABEL_FONT
        ws5[f"{c}4"].fill = PatternFill("solid", fgColor="E8EEF4")

    for i, r in enumerate(wacc_rows):
        formula_cell(ws5, f"A{r}", f"=WACC!$B$21+({wacc_offsets[i]})", PCT_FMT)
        ws5[f"A{r}"].font = BOLD_LABEL_FONT
        ws5[f"A{r}"].fill = PatternFill("solid", fgColor="E8EEF4")

        for c in growth_cols:
            formula = (
                f"=(SUMPRODUCT(Forecast!$B$11:$F$11,1/(1+$A{r})^{{1,2,3,4,5}})"
                f"+((Forecast!$F$11*(1+{c}$4))/($A{r}-{c}$4))/(1+$A{r})^Assumptions!$B$19"
                f"-Assumptions!$B$9)/Assumptions!$B$10"
            )
            formula_cell(ws5, f"{c}{r}", formula, CURRENCY_FMT)

    ws5["A11"] = "Note: cells where WACC <= Growth will show an error - not a valid combination."
    ws5["A11"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")
    ws5.merge_cells("A11:F11")

    # ================================================================
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    build_workbook()
