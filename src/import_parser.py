"""
import_parser.py
-------------------
Smart importer for uploaded files. Auto-detects which of two formats
the user uploaded and extracts usable DCF assumptions either way:

    1. "Historical" format  -> raw year-by-year financials
       (year, revenue, ebit, tax, depreciation, capex, net_debt,
        shares_outstanding, current_price columns)

    2. "Prebuilt Values" format -> an already-computed DCF workbook
       (e.g. a "DCF Values" sheet with Revenue/EBIT/FCFF rows per
       forecast year, and a "Key Assumptions" sheet with WACC,
       Terminal Growth, Tax Rate, Net Debt, Shares Outstanding,
       Current Market Price). This is common when someone hands you
       a finished model rather than raw financial statements.

For format 2, this module reverse-engineers the ratios (average
revenue growth, EBIT margin, D&A %, Capex %, NWC %) so the numbers
can be fed straight back into DCFModel and reproduce a consistent
valuation.
"""

import re
import openpyxl
import pandas as pd

HISTORICAL_REQUIRED_COLUMNS = [
    "year", "revenue", "ebit", "tax", "depreciation",
    "capex", "net_debt", "shares_outstanding", "current_price"
]

# Row labels we look for inside a "values" style sheet (case-insensitive, partial match ok)
VALUE_ROW_KEYS = {
    "revenue": ["revenue"],
    "revenue_growth": ["revenue growth"],
    "ebit_margin": ["ebit margin"],
    "ebit": ["ebit"],
    "tax_rate": ["tax rate"],
    "da": ["depreciation", "d&a", "d & a"],
    "capex": ["capital expenditure", "capex"],
    "nwc": ["net working capital", "nwc"],
    "fcff": ["fcff"],
}

ASSUMPTION_KEY_MAP = {
    "wacc": ["wacc"],
    "terminal_growth": ["terminal growth"],
    "tax_rate": ["tax rate"],
    "net_debt": ["net debt"],
    "shares_outstanding": ["shares outstanding"],
    "current_price": ["current market price", "current price", "market price"],
}


def _match_key(label: str, keywords: list) -> bool:
    label_l = label.lower()
    return any(kw in label_l for kw in keywords)


def _clean_label(cell_value):
    if cell_value is None:
        return None
    return re.sub(r"\s+", " ", str(cell_value)).strip()


def detect_historical_format(df: pd.DataFrame) -> bool:
    """Returns True if the dataframe already matches our raw historical schema."""
    cols = [c.strip().lower() for c in df.columns.astype(str)]
    return all(c in cols for c in HISTORICAL_REQUIRED_COLUMNS)


def parse_prebuilt_workbook(file_obj) -> dict:
    """
    Scans every sheet of an uploaded workbook looking for:
      - a "values" sheet with rows like Revenue / EBIT / Tax Rate / FCFF
      - an "assumptions" sheet with Assumption/Value style rows

    Returns a dict:
        {
            "found": bool,
            "params": {...}    # ready to feed into DCFModel, only keys that were found
            "raw_assumptions": {...},
            "years_detected": int,
        }
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)

    row_series = {}   # key -> list of numeric values found across forecast years
    assumptions = {}  # free-form label -> value, from an Assumption/Value sheet

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [_clean_label(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

        is_assumption_sheet = header and any(
            h and h.lower() == "assumption" for h in header
        )

        for row in ws.iter_rows(min_row=2 if is_assumption_sheet else 1):
            label = _clean_label(row[0].value)
            if not label:
                continue

            values = [c.value for c in row[1:] if isinstance(c.value, (int, float))]
            if not values:
                continue

            if is_assumption_sheet:
                assumptions[label] = values[0]
                continue

            for key, keywords in VALUE_ROW_KEYS.items():
                if _match_key(label, keywords):
                    row_series.setdefault(key, values)

    if not row_series and not assumptions:
        return {"found": False, "params": {}, "raw_assumptions": {}, "years_detected": 0}

    params = {}

    # --- Base revenue + growth rate ---
    if "revenue" in row_series and len(row_series["revenue"]) >= 1:
        revenues = row_series["revenue"]
        params["revenue_base"] = revenues[0]
        if "revenue_growth" in row_series and row_series["revenue_growth"]:
            params["revenue_growth"] = sum(row_series["revenue_growth"]) / len(row_series["revenue_growth"])
        elif len(revenues) >= 2:
            growths = [(revenues[i] / revenues[i - 1]) - 1 for i in range(1, len(revenues))]
            params["revenue_growth"] = sum(growths) / len(growths)

    # --- EBIT margin ---
    if "ebit_margin" in row_series and row_series["ebit_margin"]:
        params["ebit_margin"] = sum(row_series["ebit_margin"]) / len(row_series["ebit_margin"])
    elif "ebit" in row_series and "revenue" in row_series:
        margins = [e / r for e, r in zip(row_series["ebit"], row_series["revenue"])]
        params["ebit_margin"] = sum(margins) / len(margins)

    # --- Tax rate ---
    if "tax_rate" in row_series and row_series["tax_rate"]:
        params["tax_rate"] = sum(row_series["tax_rate"]) / len(row_series["tax_rate"])

    # --- D&A / Capex / NWC as % of revenue ---
    if "revenue" in row_series:
        revs = row_series["revenue"]
        for key, param_name in [("da", "da_pct"), ("capex", "capex_pct"), ("nwc", "nwc_pct")]:
            if key in row_series and len(row_series[key]) == len(revs):
                ratios = [v / r for v, r in zip(row_series[key], revs)]
                params[param_name] = sum(ratios) / len(ratios)

    if "revenue" in row_series:
        params["forecast_years"] = len(row_series["revenue"])

    # --- Assumptions sheet (WACC, terminal growth, net debt, shares, price) ---
    for param_key, keywords in ASSUMPTION_KEY_MAP.items():
        for label, value in assumptions.items():
            if _match_key(label, keywords):
                params[param_key] = value
                break

    return {
        "found": True,
        "params": params,
        "raw_assumptions": assumptions,
        "years_detected": params.get("forecast_years", 0),
    }


def smart_load(uploaded_file) -> dict:
    """
    Main entry point for the dashboard. Given a Streamlit UploadedFile
    (csv or xlsx), figures out which format it is and returns:

        {
            "mode": "historical" | "prebuilt" | "unrecognized",
            "df": DataFrame or None,          # for historical mode
            "params": dict,                    # for prebuilt mode
            "raw_assumptions": dict,
        }
    """
    import os
    file_ext = os.path.splitext(uploaded_file.name)[1].lower()

    # First, try the simple historical schema (CSV or Excel)
    try:
        uploaded_file.seek(0)
        if file_ext == ".csv":
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file, engine="openpyxl")

        if detect_historical_format(df):
            df.columns = [c.strip().lower() for c in df.columns.astype(str)]
            df = df.sort_values("year").reset_index(drop=True)
            return {"mode": "historical", "df": df, "params": {}, "raw_assumptions": {}}
    except Exception:
        pass

    # Not a historical CSV — if it's Excel, try the "prebuilt values" parser
    if file_ext in (".xlsx", ".xls"):
        try:
            uploaded_file.seek(0)
            result = parse_prebuilt_workbook(uploaded_file)
            if result["found"] and result["params"]:
                return {
                    "mode": "prebuilt",
                    "df": None,
                    "params": result["params"],
                    "raw_assumptions": result["raw_assumptions"],
                }
        except Exception:
            pass

    return {"mode": "unrecognized", "df": None, "params": {}, "raw_assumptions": {}}
